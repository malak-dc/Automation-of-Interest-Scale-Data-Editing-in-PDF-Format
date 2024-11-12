import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import sys
import mysql.connector
import pandas as pd
import os
import win32com.client

# Vérifiez que suffisamment d'arguments ont été fournis
if len(sys.argv) < 8:
    print("Usage: generate_excel.py <numero_compte> <date_arrete1> <date_arrete> <num_compte> <cd_agence> <intitule_compte> <lib_agence>")
    sys.exit(1)

# Récupérer les arguments passés par la ligne de commande
numero_compte = sys.argv[1]
date_arrete1 = sys.argv[2]
date_arrete = sys.argv[3]
num_compte = sys.argv[4]
cd_agence = sys.argv[5]
intitule_compte = sys.argv[6]
lib_agence = sys.argv[7]

# Affichage des arguments pour le débogage
print(f"numero_compte: {numero_compte}")
print(f"date_arrete1: {date_arrete1}")
print(f"date_arrete: {date_arrete}")
print(f"num_compte: {num_compte}")
print(f"cd_agence: {cd_agence}")
print(f"intitule_compte: {intitule_compte}")
print(f"lib_agence: {lib_agence}")

# Configuration de la connexion à la base de données
config = {
    'user': 'root',
    'password': 'admin',
    'host': 'localhost',
    'database': 'tc_prod'
}

# Connexion à la base de données
conn = mysql.connector.connect(**config)
cursor = conn.cursor()

# Requête SQL pour récupérer les données de mouvement
query_mouvement = """
SELECT
    T.DATE_VALEUR,
    T.DATE_OPERATION,
    T.LIBELLE,
    CASE
        WHEN T.MONTANT >= 0 THEN T.MONTANT
        ELSE NULL
    END AS DEBIT,
    CASE
        WHEN T.MONTANT < 0 THEN -T.MONTANT
        ELSE NULL
    END AS CREDIT,
    CASE
        WHEN T.SOLDE >= 0 THEN T.SOLDE
        ELSE NULL
    END AS SLD_DEBIT,
    CASE
        WHEN T.SOLDE < 0 THEN -T.SOLDE
        ELSE NULL
    END AS SLD_CREDIT,
    T.NBR_JOURS,
    CASE
        WHEN T.NOMBRES >= 0 THEN T.NOMBRES
        ELSE NULL
    END AS NOMBRES_DB,
    CASE
        WHEN T.NOMBRES < 0 THEN -T.NOMBRES
        ELSE NULL
    END AS NOMBRES_CR,
    T.OBSERVATION
FROM
    T_MOUVEMENT T
WHERE
    T.NUMERO_COMPTE = %s
    AND T.DATE_ARRETE = %s
ORDER BY
    T.IDENTIFIANT ASC
"""

# Requête SQL pour récupérer les taux spécifiques au compte et à la date
query_taux = """
SELECT
    tauxDebit,
    tauxCredit
FROM
    production_agios_1112
WHERE
    numeroCompte = %s
    AND dateArrete = %s
LIMIT 1
"""

# Exécution des requêtes pour les mouvements et les taux
cursor.execute(query_mouvement, (numero_compte, date_arrete))
rows = cursor.fetchall()
columns = [desc[0] for desc in cursor.description]

cursor.execute(query_taux, (numero_compte, date_arrete))
taux = cursor.fetchone()

# Fermeture de la connexion
cursor.close()
conn.close()

# Création d'un DataFrame avec les résultats de la requête
df = pd.DataFrame(rows, columns=columns)

# Création du fichier Excel
wb = Workbook()
ws = wb.active

# Remplissage gris pour certaines lignes
gray_fill = PatternFill("solid", fgColor="D3D3D3")

# Fusionner les cellules pour l'arrêté de compte et le nom
ws.merge_cells('A7:K7')
ws['A7'] = f"ARRETE DE COMPTE DU: {date_arrete1} AU: {date_arrete}"
ws['A7'].alignment = Alignment(horizontal='center', vertical='center')
ws['A7'].font = Font(size=10, bold=True)
ws['A7'].fill = gray_fill

# Fusionner les cellules pour l'agence
ws.merge_cells('A8:K8')
ws['A8'] = f"AGENCE : {lib_agence}"
ws['A8'].alignment = Alignment(horizontal='left', vertical='center')
ws['A8'].font = Font(size=10, bold=True)
ws['A8'].fill = gray_fill

# Fusionner les cellules pour le numéro de compte
ws.merge_cells('A9:F9')
ws['A9'] = f"NUMERO DE COMPTE : {numero_compte}"
ws['A9'].alignment = Alignment(horizontal='left', vertical='center')
ws['A9'].font = Font(size=10, bold=True)
ws['A9'].fill = gray_fill

# Fusionner les cellules pour le nom du compte
ws.merge_cells('G9:K9')
ws['G9'] = f"NOM : {intitule_compte}"
ws['G9'].alignment = Alignment(horizontal='left', vertical='center')
ws['G9'].font = Font(size=10, bold=True)
ws['G9'].fill = gray_fill

# Lignes vides avant les titres de colonnes
ws.append([])  # Ligne 10
ws.append([])  # Ligne 11

# Titre des colonnes à partir de la ligne 12
columns = ["DATE VALEUR", "DATE OPERAT", "LIBELLE MOUVEMENT", "MOUVEMENTS", "", "SOLDE EN VALEUR", "", "NBRE DE JOUR", "NOMBRES", "", "OBSERVATIONS"]
sub_columns = ["", "", "", "DEBIT", "CREDIT", "DEBIT", "CREDIT", "", "DEBIT", "CREDIT", ""]

# Ajouter les titres des colonnes à la feuille de calcul
ws.append(columns)  # Ligne 12
ws.append(sub_columns)  # Ligne 13

# Fusionner les cellules des en-têtes à partir de la ligne 12
ws.merge_cells('A12:A13')  # DATE VALEUR
ws.merge_cells('B12:B13')  # DATE OPERAT
ws.merge_cells('C12:C13')  # LIBELLE MOUVEMENT
ws.merge_cells('H12:H13')  # NBRE DE JOUR
ws.merge_cells('K12:K13')  # OBSERVATIONS
ws.merge_cells('D12:E12')  # MOUVEMENTS
ws.merge_cells('F12:G12')  # SOLDE EN VALEUR
ws.merge_cells('I12:J12')  # NOMBRES

# Styles des titres des colonnes
header_font = Font(bold=True, size=10)
center_alignment = Alignment(horizontal='center', vertical='center')

# Centrer les noms des colonnes
for row in ws.iter_rows(min_row=12, max_row=13, min_col=1, max_col=11):
    for cell in row:
        cell.font = header_font
        cell.alignment = center_alignment

# Ajouter les données à la feuille de calcul
start_row = 14  # Ajustement pour tenir compte de la ligne vide ajoutée
for row in rows:
    formatted_row = [
        row[0],  # DATE_VALEUR
        row[1],  # DATE_OPERATION
        row[2],  # LIBELLE
        row[3],  # DEBIT
        row[4],  # CREDIT
        row[5],  # SLD_DEBIT
        row[6],  # SLD_CREDIT
        row[7],  # NBR_JOURS
        row[8],  # NOMBRES_DB
        row[9],  # NOMBRES_CR
        row[10]  # OBSERVATION
    ]
    ws.append(formatted_row)

# Insérer une ligne vide avant les totaux
ws.append([])

# Calculer les totaux pour DEBIT et CREDIT
last_data_row = start_row + len(rows) - 1
total_row = last_data_row + 2
ws.cell(row=total_row, column=1, value="TOTAL").font = header_font
ws.cell(row=total_row, column=4, value=f"=SUM(D{start_row}:D{last_data_row})").font = header_font  # Total DEBIT
ws.cell(row=total_row, column=5, value=f"=SUM(E{start_row}:E{last_data_row})").font = header_font  # Total CREDIT
ws.cell(row=total_row, column=6, value=f"=SUM(F{start_row}:F{last_data_row})").font = header_font  # Total SLD_DEBIT
ws.cell(row=total_row, column=7, value=f"=SUM(G{start_row}:G{last_data_row})").font = header_font  # Total SLD_CREDIT
ws.cell(row=total_row, column=9, value=f"=SUM(I{start_row}:I{last_data_row})").font = header_font  # Total NOMBRES_DB
ws.cell(row=total_row, column=10, value=f"=SUM(J{start_row}:J{last_data_row})").font = header_font  # Total NOMBRES_CR

# Solde en valeur au date arrete
ws.cell(row=total_row+3, column=1, value=f"SOLDE EN VALEUR AU {date_arrete}").font = header_font
ws.cell(row=total_row+3, column=3, value=f"{rows[-1][5]} DH").font = header_font  # Dernière valeur du solde en valeur débit
# Exemple de ligne contenant "TOTAL"
total_row = last_data_row + 2
total_column = 1  # Colonne où le texte "TOTAL" est placé

# Ajouter le texte "TOTAL" et centrer
ws.cell(row=total_row, column=total_column, value="TOTAL").alignment = Alignment(horizontal='center', vertical='center')
ws.cell(row=total_row, column=total_column).font = Font(bold=True)
# Calcul au taux
if taux:
    taux_debit, taux_credit = taux
    ws.cell(row=total_row+5, column=1, value="CALCUL AU TAUX : ").font = header_font
    ws.cell(row=total_row+5, column=3, value=f"DB {taux_debit}")
    ws.cell(row=total_row+5, column=4, value=f"CR {taux_credit}")

    # Calcul des intérêts
    for i in range(start_row, last_data_row+1):
        ws.cell(row=i, column=12, value=f"=I{i}*{taux_debit}/360")  # Intérêt DEBIT
        ws.cell(row=i, column=13, value=f"=J{i}*{taux_credit}/360")  # Intérêt CREDIT
else:
    print(f"Erreur : Les taux pour le compte {numero_compte} et la date {date_arrete} n'ont pas été trouvés.")


# Appliquer des bordures à toutes les cellules du tableau
thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

for row in ws.iter_rows(min_row=12, max_row=last_data_row, min_col=1, max_col=11):
    for cell in row:
        cell.border = thin_border
for row in ws.iter_rows(min_row=last_data_row, max_row=last_data_row+2, min_col=1, max_col=11):
    for cell in row:
        cell.border = thin_border

# Ajuster la largeur des colonnes
for column_cells in ws.iter_cols():
    length = max(len(str(cell.value)) for cell in column_cells)
    column_letter = get_column_letter(column_cells[0].column)
    ws.column_dimensions[column_letter].width = min(length + 2, 20)

# Réduire la hauteur des lignes
for row in ws.iter_rows():
    ws.row_dimensions[row[0].row].height = 12

# Réduire la largeur de la deuxième colonne (Date Operat)
ws.column_dimensions['B'].width = 12

# Appliquer le remplissage gris aux lignes spécifiques
for row in [total_row+3, total_row+4, total_row+5]:
    for col in range(1, 12):  # De la colonne A à K (1 à 11 en index)
        cell = ws.cell(row=row, column=col)
        cell.fill = gray_fill

# Variables pour les coordonnées de fusion
start_row = last_data_row + 1  # Définir la ligne de début après les dernières données
end_row = start_row  # Même ligne si vous ne voulez fusionner qu'une seule ligne horizontalement
start_col = 1  # Par exemple, colonne A
end_col = 11  # Par exemple, colonne K (11ème colonne)

# Fusionner les cellules de la plage spécifiée
ws.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)
# Variables pour les coordonnées de fusion
start_row = last_data_row + 2  # Définir la ligne de début après les dernières données
end_row = start_row  # Même ligne si vous ne voulez fusionner qu'une seule ligne horizontalement
start_col = 1  # Par exemple, colonne A
end_col = 3  # Par exemple, colonne K (11ème colonne)

# Fusionner les cellules de la plage spécifiée
ws.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)


# Sauvegarde du fichier Excel
file_path = f"CANVA_{numero_compte}.xlsx"
wb.save(file_path)


# Convertir Excel en PDF
try:
    excel = win32com.client.Dispatch("Excel.Application")
    workbook = excel.Workbooks.Open(os.path.abspath(file_path))

    # Fit to page
    workbook.Sheets(1).PageSetup.Zoom = False
    workbook.Sheets(1).PageSetup.FitToPagesWide = 1
    workbook.Sheets(1).PageSetup.FitToPagesTall = False

    pdf_path = os.path.abspath(file_path.replace(".xlsx", ".pdf"))
    workbook.ExportAsFixedFormat(0, pdf_path)
    workbook.Close(False)
    excel.Quit()

    # Vérifiez si le fichier PDF a été créé avec succès
    if os.path.exists(pdf_path):
        print(f"Fichier PDF généré : {pdf_path}")
        os.startfile(pdf_path)  # Ouvrir le PDF automatiquement
    else:
        print(f"Erreur : Le fichier PDF {pdf_path} n'a pas été créé.")
except Exception as e:
    print(f"Erreur lors de la conversion en PDF : {e}")